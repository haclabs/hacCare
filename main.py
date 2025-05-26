"""
Author: haclabs
Application: hacCare Record Scanner (Windows GUI)
Version: 1.4.4
Date: 2025-05-20

Description:
This is the Windows GUI version of hacCare with support for tkinter and winsound.
"""

import tkinter as tk
from tkinter import simpledialog, messagebox, filedialog
import openpyxl
import os
import webbrowser
import shutil
import winsound

EXCEL_FILE = "records.xlsx"
FOLDERS = {"Record": "Records", "Med": "Meds"}
"""
Author: haclabs
Application: hacCare Record Scanner (Windows GUI)
Version: 1.4.4 (with auto-generated HTML patient record)
Date: 2025-05-20

Description:
Windows GUI version of hacCare with support for tkinter and winsound.
Auto-generates a sample patient HTML record if no file is uploaded for 'Record' type.
"""

import tkinter as tk
from tkinter import simpledialog, messagebox, filedialog
import openpyxl
import os
import webbrowser
import shutil
import winsound

EXCEL_FILE = "records.xlsx"
FOLDERS = {"Record": "Records", "Med": "Meds"}

def play_success_beep():
    winsound.Beep(1000, 150)

def play_error_beep():
    winsound.Beep(300, 300)

def load_records():
    records = {}
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Number", "Type", "File"])
        wb.save(EXCEL_FILE)
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            number = str(row[0])
            folder = FOLDERS.get(row[1], "")
            path = os.path.join(folder, row[2])
            records[number] = path
    return records

def save_record_to_excel(number, record_type, filename):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([number, record_type, filename])
    wb.save(EXCEL_FILE)

def open_record():
    number = entry.get().strip()
    if number in records:
        record = records[number]
        if os.path.exists(record):
            os.startfile(record)
            play_success_beep()
        else:
            messagebox.showerror("Missing File", f"File not found: {record}")
            play_error_beep()
    else:
        messagebox.showwarning("Not Found", "Record number not found.")
        play_error_beep()
    entry.delete(0, tk.END)

def add_record():
    number = simpledialog.askstring("Input", "Enter record number:")
    if not number:
        return

    type_choice = simpledialog.askstring("Input", "Enter type (Record or Med):", initialvalue="Record")
    if type_choice not in FOLDERS:
        messagebox.showerror("Invalid Type", "Please enter 'Record' or 'Med'.")
        return

    if type_choice == "Record":
        # Ask if user wants to auto-generate or upload
        choice = messagebox.askyesno("Sample Record", "Generate a sample hospital patient record HTML?\n\nClick 'No' to upload your own file.")
        if choice:
            folder = FOLDERS[type_choice]
            os.makedirs(folder, exist_ok=True)
            filename = f"patient_record_{number}.html"
            dest_path = os.path.join(folder, filename)
            html_content = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Patient Record - Sample</title>
    <style>
        body { font-family: Arial, sans-serif; background: #f5f7fa; margin: 40px; }
        .record-container { background: #fff; padding: 32px; border-radius: 14px; box-shadow: 0 2px 12px #0001; max-width: 600px; margin: auto; }
        h1 { color: #2b3553; }
        table { width: 100%; border-collapse: collapse; margin-top: 20px; }
        th, td { text-align: left; padding: 8px 12px; border-bottom: 1px solid #eee; }
        th { background: #f0f3fa; }
    </style>
</head>
<body>
    <div class="record-container">
        <h1>Hospital Patient Record</h1>
        <table>
            <tr><th>Patient Name</th><td>Jane Doe</td></tr>
            <tr><th>Patient ID</th><td>123456</td></tr>
            <tr><th>Date of Birth</th><td>1991-03-14</td></tr>
            <tr><th>Gender</th><td>Female</td></tr>
            <tr><th>Admission Date</th><td>2025-05-20</td></tr>
            <tr><th>Diagnosis</th><td>Pneumonia</td></tr>
            <tr><th>Allergies</th><td>Penicillin</td></tr>
            <tr><th>Medications</th><td>Amoxicillin, Acetaminophen</td></tr>
            <tr><th>Attending Physician</th><td>Dr. John Smith</td></tr>
        </table>
        <p style="margin-top:24px;"><strong>Notes:</strong> <br>This is a sample patient record for demonstration purposes.</p>
    </div>
</body>
</html>"""
            with open(dest_path, "w", encoding="utf-8") as f:
                f.write(html_content)
            records[number] = dest_path
            save_record_to_excel(number, type_choice, filename)
            messagebox.showinfo("Success", f"Sample patient record HTML generated as {filename}.")
        else:
            file_path = filedialog.askopenfilename(title="Select File")
            if not file_path:
                return
            folder = FOLDERS[type_choice]
            os.makedirs(folder, exist_ok=True)
            filename = os.path.basename(file_path)
            dest_path = os.path.join(folder, filename)
            shutil.copy(file_path, dest_path)
            records[number] = dest_path
            save_record_to_excel(number, type_choice, filename)
            messagebox.showinfo("Success", "Record added.")
    else:
        file_path = filedialog.askopenfilename(title="Select File")
        if not file_path:
            return
        folder = FOLDERS[type_choice]
        os.makedirs(folder, exist_ok=True)
        filename = os.path.basename(file_path)
        dest_path = os.path.join(folder, filename)
        shutil.copy(file_path, dest_path)
        records[number] = dest_path
        save_record_to_excel(number, type_choice, filename)
        messagebox.showinfo("Success", "Record added.")

root = tk.Tk()
root.title("hacCare v1.4.4")
root.geometry("400x300")

records = load_records()

tk.Label(root, text="Welcome to hacCare", font=("Helvetica", 16)).pack(pady=10)

entry = tk.Entry(root, font=("Arial", 14))
entry.pack(pady=10)
entry.bind("<Return>", lambda e: open_record())

tk.Button(root, text="Open Record", command=open_record).pack(pady=5)
tk.Button(root, text="Add New Record", command=add_record).pack(pady=5)
tk.Button(root, text="Exit", command=root.quit).pack(pady=5)

root.mainloop()

def play_success_beep():
    winsound.Beep(1000, 150)

def play_error_beep():
    winsound.Beep(300, 300)

def load_records():
    records = {}
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Number", "Type", "File"])
        wb.save(EXCEL_FILE)
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            number = str(row[0])
            folder = FOLDERS.get(row[1], "")
            path = os.path.join(folder, row[2])
            records[number] = path
    return records

def save_record_to_excel(number, record_type, filename):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([number, record_type, filename])
    wb.save(EXCEL_FILE)

def open_record():
    number = entry.get().strip()
    if number in records:
        record = records[number]
        if os.path.exists(record):
            os.startfile(record)
            play_success_beep()
        else:
            messagebox.showerror("Missing File", f"File not found: {record}")
            play_error_beep()
    else:
        messagebox.showwarning("Not Found", "Record number not found.")
        play_error_beep()
    entry.delete(0, tk.END)

def add_record():
    number = simpledialog.askstring("Input", "Enter record number:")
    if not number:
        return

    type_choice = simpledialog.askstring("Input", "Enter type (Record or Med):", initialvalue="Record")
    if type_choice not in FOLDERS:
        messagebox.showerror("Invalid Type", "Please enter 'Record' or 'Med'.")
        return

    file_path = filedialog.askopenfilename(title="Select File")
    if not file_path:
        return

    folder = FOLDERS[type_choice]
    os.makedirs(folder, exist_ok=True)
    filename = os.path.basename(file_path)
    dest_path = os.path.join(folder, filename)
    shutil.copy(file_path, dest_path)
    records[number] = dest_path
    save_record_to_excel(number, type_choice, filename)
    messagebox.showinfo("Success", "Record added.")

root = tk.Tk()
root.title("hacCare v1.4.3")
root.geometry("400x300")

records = load_records()

tk.Label(root, text="Welcome to hacCare", font=("Helvetica", 16)).pack(pady=10)

entry = tk.Entry(root, font=("Arial", 14))
entry.pack(pady=10)
entry.bind("<Return>", lambda e: open_record())

tk.Button(root, text="Open Record", command=open_record).pack(pady=5)
tk.Button(root, text="Add New Record", command=add_record).pack(pady=5)
tk.Button(root, text="Exit", command=root.quit).pack(pady=5)

root.mainloop()
